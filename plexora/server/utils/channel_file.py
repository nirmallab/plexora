# Reading a list of channel names out of whatever file the user happens to
# have. Backs POST /upload_channels (see server/routes/data_routes.py).
#
# The file is nearly always a one-column list exported from a panel design
# sheet -- one marker per row, in the order the image stacks them -- and that
# case has to work with no questions asked. Everything else in here exists for
# the files that are not that: a spreadsheet with the names in the third
# column, a header row that may or may not be one, a table that turns out to
# belong to a different image entirely.
#
# Parsing is server-side for all of them, including the browser upload, because
# the OTHER way in is a path typed into the box for it: on a cluster the file
# sits beside the image on the remote filesystem, where the browser cannot
# reach it and only the server can. One reader rather than two means the path
# and the upload cannot disagree about what a file says.

import csv
import io
from pathlib import Path

#: What the picker shows of the table. Five and five is enough to recognise a
#: file by; anything past it is reached by scrolling the preview sideways
#: rather than by making the modal bigger.
PREVIEW_ROWS = 5
PREVIEW_COLUMNS = 5

#: How many columns the picker's select may offer. A channel list has one or
#: two; this is the guard against someone reaching for a feature table with
#: four hundred, which would be a select nobody can use.
MAX_COLUMNS = 64

#: Read no further than this. A channel list is as long as the image is deep --
#: tens of rows, a few hundred at the very outside -- so a file past this is
#: not one, and reading a million-row table into memory to tell the user so is
#: a slow way to reach an answer that was available at row 20,001.
MAX_ROWS = 20000

DELIMITED_SUFFIXES = (".csv", ".tsv", ".txt")
EXCEL_SUFFIXES = (".xlsx", ".xlsm")
#: Read by neither openpyxl nor the csv module. Named separately so the message
#: can say what to do about it rather than listing what is accepted.
LEGACY_EXCEL_SUFFIXES = (".xls",)

#: One sentence, used by the route's rejection message and nothing else. Kept
#: here so the reader and the refusal cannot drift apart.
SUPPORTED_DESCRIPTION = "a .csv, .tsv, .txt, .xlsx or .xlsm file"


class ChannelFileError(ValueError):
    """Something the user can act on, already worded for them.

    Every message reaching this class is shown verbatim in the browser, so it
    names the file or the fix rather than the function that gave up.
    """


def _suffix(filename):
    return Path(filename or "").suffix.lower()


def _sniff_delimiter(text, suffix):
    """Which character separates the cells.

    A .tsv says so in its name. For everything else this counts candidates in
    the first non-empty line rather than using csv.Sniffer, which reads a
    sample and guesses at quoting too -- and on a one-column file of bare
    marker names has nothing to go on and raises. One column is the common
    case here, and the comma it falls back to is right for it either way.
    """
    if suffix == ".tsv":
        return "\t"
    for line in text.splitlines():
        if not line.strip():
            continue
        counts = {candidate: line.count(candidate) for candidate in (",", "\t", ";", "|")}
        best = max(counts, key=counts.get)
        return best if counts[best] else ","
    return ","


def _read_delimited(data, suffix):
    # utf-8-sig: a sheet saved as CSV from Excel on Windows leads with a BOM,
    # which would otherwise ride along on the first channel's name and match
    # nothing. errors="replace" rather than a second decode attempt -- a name
    # with one unreadable byte in it is still a name the user can recognise
    # and correct, and a hard failure here would refuse the whole file.
    text = data.decode("utf-8-sig", errors="replace")
    delimiter = _sniff_delimiter(text, suffix)
    rows = []
    for row in csv.reader(io.StringIO(text), delimiter=delimiter):
        rows.append(row)
        if len(rows) > MAX_ROWS:
            raise ChannelFileError(_too_long())
    return rows


def _read_excel(data):
    try:
        from openpyxl import load_workbook
    except ImportError:
        # Declared in pyproject, so this is an environment that has drifted
        # rather than an unsupported format. Say which of the two it is.
        raise ChannelFileError(
            "Reading Excel files needs the openpyxl package, which is not installed "
            "in this environment. Install it, or save the sheet as CSV."
        ) from None

    try:
        # read_only streams the sheet a row at a time instead of building the
        # whole cell graph, and data_only hands back the last computed value of
        # a formula rather than the formula text -- a name column produced by
        # =CONCAT(...) is otherwise a column of "=CONCAT(...)".
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ChannelFileError(f"That file could not be opened as a spreadsheet ({exc}).") from None

    try:
        # The sheet the file was saved on. A workbook of several sheets is a
        # real possibility and there is nothing here that could pick between
        # them, so the one the author left showing is the best available guess
        # -- and a wrong guess lands in the column picker, not in the project.
        sheet = workbook.active
        if sheet is None:
            raise ChannelFileError("That spreadsheet has no sheets in it.")
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) > MAX_ROWS:
                raise ChannelFileError(_too_long())
        return rows
    finally:
        workbook.close()


def _too_long():
    return (
        f"That file has more than {MAX_ROWS:,} rows, so it is not a list of channel "
        "names. Pick the file that has one name per channel."
    )


def _normalize(rows):
    """Rows of raw cells into a rectangle of stripped strings.

    Blank rows go: a trailing newline, a spacer line between sections, and the
    empty rows Excel reports past the end of the data are all noise, and each
    of them would otherwise count as a channel the image does not have. Empty
    trailing COLUMNS go for the same reason -- a sheet where somebody once
    typed in column H reports eight columns forever, and the picker would offer
    five empty ones.
    """
    grid = []
    for row in rows:
        cells = [("" if cell is None else str(cell)).strip() for cell in row]
        if any(cells):
            grid.append(cells)

    width = max((len(row) for row in grid), default=0)
    grid = [row + [""] * (width - len(row)) for row in grid]
    while width and not any(row[width - 1] for row in grid):
        width -= 1
        grid = [row[:width] for row in grid]
    return grid


def read_grid(data=None, path=None, filename=None):
    """The file as a rectangle of stripped strings.

    Exactly one of `data` (bytes, from a browser upload) or `path` (a file on
    the machine running the server) is given; `filename` decides the format and
    defaults to the path's own name.
    """
    if filename is None:
        filename = Path(path).name if path else ""
    suffix = _suffix(filename)

    if suffix in LEGACY_EXCEL_SUFFIXES:
        raise ChannelFileError(
            "Old-style .xls files are not supported. Open it and save it as .xlsx or .csv."
        )
    if suffix not in DELIMITED_SUFFIXES and suffix not in EXCEL_SUFFIXES:
        raise ChannelFileError(
            f"{filename or 'That file'} is not {SUPPORTED_DESCRIPTION}."
        )

    if data is None:
        try:
            data = Path(path).read_bytes()
        except OSError as exc:
            raise ChannelFileError(f"That file could not be read ({exc.strerror or exc}).") from None

    rows = _read_excel(data) if suffix in EXCEL_SUFFIXES else _read_delimited(data, suffix)
    grid = _normalize(rows)
    if not grid:
        raise ChannelFileError(f"{filename or 'That file'} is empty.")
    return grid


def width(grid):
    return len(grid[0]) if grid else 0


def names(grid, column, has_header):
    """The channel names one column holds, in file order.

    Empty cells are dropped rather than carried through as unnamed channels:
    in a multi-column table the name column is routinely shorter than the
    others, and a blank is not a channel called "".
    """
    if column < 0 or column >= width(grid):
        raise ChannelFileError("That column is not in the file.")
    rows = grid[1:] if has_header else grid
    return [row[column] for row in rows if row[column]]


def autodetect(grid, channel_count):
    """Whether this file answers the question on its own, and how.

    Only a single-column file does. Two readings are tried, in this order:
    every row is a name, or every row but the first is and the first is the
    column's title. Anything else -- more than one column, or a count that
    matches neither -- is a question for the user, and returns None rather
    than a reading that would rename the image's channels to the wrong list.

    @returns True/False (the file has a header row / it does not), or None.
    """
    if width(grid) != 1:
        return None
    for has_header in (False, True):
        if len(names(grid, 0, has_header)) == channel_count:
            return has_header
    return None


def describe(grid, channel_count, filename=""):
    """What the column picker needs to draw itself.

    Small on purpose -- five rows of the table, and one entry per column
    carrying the count the user is actually comparing against the image. The
    counts are what let the modal keep its "N names / M channels" line honest
    as the header checkbox is ticked, without asking the server again.
    """
    columns = []
    for index in range(min(width(grid), MAX_COLUMNS)):
        columns.append({
            "index": index,
            "header": grid[0][index],
            "nonempty": sum(1 for row in grid if row[index]),
        })

    # A guess at the checkbox's starting state, not a claim. If dropping the
    # first row is what makes some column come out at exactly the channel
    # count, that row is a header -- which is the same reasoning autodetect
    # uses on a single column, applied to a table where the user still has to
    # say which column they mean.
    header_guess = any(
        column["nonempty"] == channel_count + 1 and grid[0][column["index"]]
        for column in columns
    )

    return {
        "filename": filename,
        "channel_count": channel_count,
        "row_count": len(grid),
        "column_count": width(grid),
        "columns": columns,
        "columns_truncated": width(grid) > MAX_COLUMNS,
        "preview": [row[:PREVIEW_COLUMNS] for row in grid[:PREVIEW_ROWS]],
        "preview_rows": PREVIEW_ROWS,
        "preview_columns": PREVIEW_COLUMNS,
        "header_guess": header_guess,
    }
